"""Version A importer: Matrixify-style linked Color & pattern variant create/update."""

from __future__ import annotations

import argparse
import csv
import sys
from collections import Counter, OrderedDict
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from swatch_linker import (
    LINKED_KEY,
    LINKED_NAMESPACE,
    Shopify,
    configured_store,
    download_workbook,
    handle_from_label,
    text,
)

INPUT_PATH = Path("input/matrixify.xlsx")
REPORT_PATH = Path("reports/import_linked_variants_results.csv")

# The only workbook fields Version A reads. All other populated fields are reported as ignored.
FIELD_ALIASES = {
    "product_id": ("ID",),
    "handle": ("Handle",),
    "variant_id": ("Variant ID",),
    "variant_command": ("Variant Command",),
    "option1_value": ("Option1 Value", "Option 1 Value"),
    "sku": ("Variant SKU",),
    "barcode": ("Variant Barcode",),
    "price": ("Variant Price",),
}

PRODUCT_DETAILS_QUERY = """
query ProductDetailsForLinkedVariantImport($id: ID!) {
  product(id: $id) {
    id
    handle
    options {
      id
      name
      position
      linkedMetafield { namespace key }
      optionValues { id name linkedMetafieldValue }
    }
    variants(first: 250) {
      nodes {
        id
        sku
        barcode
        price
        selectedOptions { name value }
      }
      pageInfo { hasNextPage }
    }
  }
}
"""

PRODUCT_BY_HANDLE_QUERY = """
query ProductDetailsByHandleForLinkedVariantImport($handle: String!) {
  productByIdentifier(identifier: {handle: $handle}) {
    id
    handle
    options {
      id
      name
      position
      linkedMetafield { namespace key }
      optionValues { id name linkedMetafieldValue }
    }
    variants(first: 250) {
      nodes {
        id
        sku
        barcode
        price
        selectedOptions { name value }
      }
      pageInfo { hasNextPage }
    }
  }
}
"""

ADD_LINKED_VALUES_MUTATION = """
mutation AddLinkedOptionValues($productId: ID!, $option: OptionUpdateInput!, $values: [OptionValueCreateInput!]) {
  productOptionUpdate(productId: $productId, option: $option, optionValuesToAdd: $values) {
    product {
      id
      options {
        id
        optionValues { id name linkedMetafieldValue }
      }
    }
    userErrors { field message code }
  }
}
"""

UPDATE_VARIANTS_MUTATION = """
mutation UpdateLinkedVariants($productId: ID!, $variants: [ProductVariantsBulkInput!]!) {
  productVariantsBulkUpdate(productId: $productId, variants: $variants, allowPartialUpdates: false) {
    productVariants { id sku barcode price selectedOptions { name value } }
    userErrors { field message code }
  }
}
"""

CREATE_VARIANTS_MUTATION = """
mutation CreateLinkedVariants($productId: ID!, $variants: [ProductVariantsBulkInput!]!) {
  productVariantsBulkCreate(productId: $productId, variants: $variants) {
    productVariants { id sku barcode price selectedOptions { name value } }
    userErrors { field message code }
  }
}
"""


@dataclass
class SourceRow:
    excel_row: int
    product_reference_type: str
    product_reference: str
    variant_id: str
    variant_command: str
    option1_value: str
    metaobject_handle: str
    sku: str
    barcode: str
    price: str
    ignored_populated_columns: set[str] = field(default_factory=set)


@dataclass
class ProductPlan:
    source_rows: list[SourceRow]
    product: dict[str, Any]
    option: dict[str, Any]
    metaobjects: dict[str, dict[str, Any]]
    add_metaobject_ids: list[str]
    update_rows: list[SourceRow]
    create_rows: list[SourceRow]
    error: str = ""


def normalized_headers(worksheet: Any) -> tuple[dict[str, int], dict[int, str]]:
    normalized: dict[str, int] = {}
    original: dict[int, str] = {}
    for column in range(1, worksheet.max_column + 1):
        header = text(worksheet.cell(1, column).value)
        if header:
            normalized[header.casefold()] = column
            original[column] = header
    return normalized, original


def column_for(headers: dict[str, int], aliases: tuple[str, ...]) -> int | None:
    for alias in aliases:
        column = headers.get(alias.casefold())
        if column:
            return column
    return None


def select_product_sheet(workbook: Any) -> Any:
    names = {name.casefold(): name for name in workbook.sheetnames}
    sheet_name = names.get("product") or names.get("products")
    if not sheet_name:
        raise RuntimeError("The Matrixify workbook needs a Product or Products sheet (case-insensitive).")
    return workbook[sheet_name]


def parse_price(value: str) -> str:
    if not value:
        return ""
    normalized = value.replace(",", "").replace("¥", "").strip()
    try:
        decimal = Decimal(normalized)
    except InvalidOperation as exc:
        raise RuntimeError(f"Variant Price {value!r} is not a number.") from exc
    if decimal < 0:
        raise RuntimeError("Variant Price cannot be negative.")
    return format(decimal, "f")


def source_groups(workbook_path: Path) -> list[list[SourceRow]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    worksheet = select_product_sheet(workbook)
    headers, original_headers = normalized_headers(worksheet)
    columns = {name: column_for(headers, aliases) for name, aliases in FIELD_ALIASES.items()}
    if not columns["option1_value"] or not (columns["product_id"] or columns["handle"]):
        raise RuntimeError("Need Option1 Value (or Option 1 Value) and either ID or Handle.")

    allowed_columns = {column for column in columns.values() if column}
    previous_reference_type = ""
    previous_reference = ""
    groups: OrderedDict[tuple[str, str], list[SourceRow]] = OrderedDict()

    for row_number in range(2, worksheet.max_row + 1):
        raw = {column: text(worksheet.cell(row_number, column).value) for column in original_headers}
        if not any(raw.values()):
            continue

        product_id = raw.get(columns["product_id"], "") if columns["product_id"] else ""
        handle = raw.get(columns["handle"], "") if columns["handle"] else ""
        if product_id:
            previous_reference_type, previous_reference = "id", product_id
        elif handle:
            previous_reference_type, previous_reference = "handle", handle

        option1_value = raw.get(columns["option1_value"], "")
        if not option1_value:
            continue
        if not previous_reference:
            raise RuntimeError(f"Excel row {row_number} needs a product ID or Handle.")

        command = raw.get(columns["variant_command"], "") if columns["variant_command"] else ""
        if command and command.casefold() != "merge":
            raise RuntimeError(
                f"Excel row {row_number} has unsupported Variant Command {command!r}; Version A accepts MERGE or blank."
            )

        ignored = {
            original_headers[column]
            for column, value in raw.items()
            if value and column not in allowed_columns
        }
        source = SourceRow(
            excel_row=row_number,
            product_reference_type=previous_reference_type,
            product_reference=previous_reference,
            variant_id=raw.get(columns["variant_id"], "") if columns["variant_id"] else "",
            variant_command=command,
            option1_value=option1_value,
            metaobject_handle=handle_from_label(option1_value),
            sku=raw.get(columns["sku"], "") if columns["sku"] else "",
            barcode=raw.get(columns["barcode"], "") if columns["barcode"] else "",
            price=parse_price(raw.get(columns["price"], "")) if columns["price"] else "",
            ignored_populated_columns=ignored,
        )
        groups.setdefault((previous_reference_type, previous_reference), []).append(source)

    if not groups:
        raise RuntimeError("No populated Option1 Value cells were found.")
    return list(groups.values())


def product_details(shopify: Shopify, reference_type: str, reference: str) -> dict[str, Any] | None:
    if reference_type == "handle":
        return shopify.graphql(PRODUCT_BY_HANDLE_QUERY, {"handle": reference}).get("productByIdentifier")
    product_id = reference if reference.startswith("gid://") else f"gid://shopify/Product/{reference}"
    return shopify.graphql(PRODUCT_DETAILS_QUERY, {"id": product_id}).get("product")


def first_option(product: dict[str, Any]) -> dict[str, Any] | None:
    options = product.get("options") or []
    return min(options, key=lambda option: option.get("position", 9999)) if options else None


def gid(resource: str, value: str) -> str:
    return value if value.startswith("gid://") else f"gid://shopify/{resource}/{value}"


def variant_option_meta_id(variant: dict[str, Any], option: dict[str, Any]) -> str:
    selected_values = {
        text(item.get("value"))
        for item in variant.get("selectedOptions") or []
        if text(item.get("name")).casefold() == text(option.get("name")).casefold()
    }
    linked_by_name = {
        text(value.get("name")): text(value.get("linkedMetafieldValue"))
        for value in option.get("optionValues") or []
    }
    for selected in selected_values:
        if selected in linked_by_name:
            return linked_by_name[selected]
    return ""


def plan_group(source_rows: list[SourceRow], shopify: Shopify) -> ProductPlan:
    source = source_rows[0]
    product = product_details(shopify, source.product_reference_type, source.product_reference)
    if not product:
        return ProductPlan(source_rows, {}, {}, {}, [], [], [], "Product not found.")
    variants_payload = product.get("variants") or {}
    if (variants_payload.get("pageInfo") or {}).get("hasNextPage"):
        return ProductPlan(source_rows, product, {}, {}, [], [], [], "Product has more than 250 variants; Version A stops safely.")
    options = product.get("options") or []
    if len(options) != 1:
        return ProductPlan(source_rows, product, {}, {}, [], [], [], "Version A supports products with exactly one option.")
    option = first_option(product)
    linked = option.get("linkedMetafield") or {}
    if (linked.get("namespace"), linked.get("key")) != (LINKED_NAMESPACE, LINKED_KEY):
        return ProductPlan(source_rows, product, option, {}, [], [], [], "First option is not linked to shopify.color-pattern.")

    metaobjects: dict[str, dict[str, Any]] = {}
    for handle in sorted({row.metaobject_handle for row in source_rows}):
        metaobject = shopify.metaobject(handle)
        if not metaobject:
            return ProductPlan(source_rows, product, option, metaobjects, [], [], [], f"Missing Color & pattern metaobject: {handle}")
        metaobjects[handle] = metaobject

    variants = variants_payload.get("nodes") or []
    variants_by_id = {variant["id"]: variant for variant in variants}
    requested_update_ids: set[str] = set()
    for row in source_rows:
        if row.variant_id:
            variant_id = gid("ProductVariant", row.variant_id)
            if variant_id not in variants_by_id:
                return ProductPlan(source_rows, product, option, metaobjects, [], [], [], f"Variant ID {row.variant_id} does not belong to this product.")
            if variant_id in requested_update_ids:
                return ProductPlan(source_rows, product, option, metaobjects, [], [], [], f"Variant ID {row.variant_id} appears more than once in the workbook.")
            requested_update_ids.add(variant_id)

    final_assignments: dict[str, str] = {
        variant["id"]: variant_option_meta_id(variant, option) for variant in variants
    }
    for row in source_rows:
        target = metaobjects[row.metaobject_handle]["id"]
        if row.variant_id:
            final_assignments[gid("ProductVariant", row.variant_id)] = target
        else:
            final_assignments[f"new:{row.excel_row}"] = target
    assigned_ids = [value for value in final_assignments.values() if value]
    duplicates = [meta_id for meta_id, count in Counter(assigned_ids).items() if count > 1]
    if duplicates:
        return ProductPlan(
            source_rows,
            product,
            option,
            metaobjects,
            [],
            [],
            [],
            "Requested variant state would duplicate linked option values: " + ", ".join(duplicates),
        )

    existing_meta_ids = {text(value.get("linkedMetafieldValue")) for value in option.get("optionValues") or []}
    add_metaobject_ids = sorted(
        {
            metaobjects[row.metaobject_handle]["id"]
            for row in source_rows
            if metaobjects[row.metaobject_handle]["id"] not in existing_meta_ids
        }
    )
    update_rows = [row for row in source_rows if row.variant_id]
    create_rows = [row for row in source_rows if not row.variant_id]
    return ProductPlan(source_rows, product, option, metaobjects, add_metaobject_ids, update_rows, create_rows)


def variant_input(row: SourceRow, option: dict[str, Any], metaobject_id: str, update: bool) -> dict[str, Any]:
    payload: dict[str, Any] = {
        "optionValues": [{"optionId": option["id"], "linkedMetafieldValue": metaobject_id}],
    }
    if update:
        payload["id"] = gid("ProductVariant", row.variant_id)
    if row.sku:
        payload["sku"] = row.sku
    if row.barcode:
        payload["barcode"] = row.barcode
    if row.price:
        payload["price"] = row.price
    return payload


def add_values(shopify: Shopify, plan: ProductPlan) -> None:
    if not plan.add_metaobject_ids:
        return
    payload = shopify.graphql(
        ADD_LINKED_VALUES_MUTATION,
        {
            "productId": plan.product["id"],
            "option": {"id": plan.option["id"]},
            "values": [{"linkedMetafieldValue": value} for value in plan.add_metaobject_ids],
        },
    ).get("productOptionUpdate") or {}
    errors = payload.get("userErrors") or []
    if errors:
        raise RuntimeError(str(errors))


def update_variants(shopify: Shopify, plan: ProductPlan) -> None:
    if not plan.update_rows:
        return
    variants = [
        variant_input(row, plan.option, plan.metaobjects[row.metaobject_handle]["id"], update=True)
        for row in plan.update_rows
    ]
    payload = shopify.graphql(
        UPDATE_VARIANTS_MUTATION,
        {"productId": plan.product["id"], "variants": variants},
    ).get("productVariantsBulkUpdate") or {}
    errors = payload.get("userErrors") or []
    if errors:
        raise RuntimeError(str(errors))


def create_variants(shopify: Shopify, plan: ProductPlan) -> None:
    if not plan.create_rows:
        return
    variants = [
        variant_input(row, plan.option, plan.metaobjects[row.metaobject_handle]["id"], update=False)
        for row in plan.create_rows
    ]
    payload = shopify.graphql(
        CREATE_VARIANTS_MUTATION,
        {"productId": plan.product["id"], "variants": variants},
    ).get("productVariantsBulkCreate") or {}
    errors = payload.get("userErrors") or []
    if errors:
        raise RuntimeError(str(errors))


def report_row(row: SourceRow, status: str, detail: str, plan: ProductPlan | None = None) -> dict[str, str]:
    return {
        "excel_row": str(row.excel_row),
        "product_reference_type": row.product_reference_type,
        "product_reference": row.product_reference,
        "variant_id": row.variant_id,
        "operation": "UPDATE" if row.variant_id else "CREATE",
        "option1_value": row.option1_value,
        "metaobject_handle": row.metaobject_handle,
        "status": status,
        "detail": detail,
        "sku": row.sku,
        "barcode": row.barcode,
        "price": row.price,
        "ignored_populated_columns": "; ".join(sorted(row.ignored_populated_columns)),
        "shopify_product_id": text(plan.product.get("id")) if plan and plan.product else "",
        "shopify_handle": text(plan.product.get("handle")) if plan and plan.product else "",
        "shopify_option_id": text(plan.option.get("id")) if plan and plan.option else "",
        "added_linked_value_ids": "; ".join(plan.add_metaobject_ids) if plan else "",
    }


def execute_plan(plan: ProductPlan, shopify: Shopify, mode: str) -> list[dict[str, str]]:
    if plan.error:
        return [report_row(row, "ERROR", plan.error, plan) for row in plan.source_rows]
    if mode == "dry-run":
        return [
            report_row(
                row,
                "WOULD_UPDATE" if row.variant_id else "WOULD_CREATE",
                "Dry run only; only allowlist fields would be sent to Shopify.",
                plan,
            )
            for row in plan.source_rows
        ]
    try:
        add_values(shopify, plan)
    except RuntimeError as exc:
        return [report_row(row, "ERROR", f"Adding linked option values failed: {exc}", plan) for row in plan.source_rows]
    try:
        update_variants(shopify, plan)
    except RuntimeError as exc:
        return [report_row(row, "ERROR", f"Updating variants failed: {exc}", plan) for row in plan.update_rows] + [
            report_row(row, "NOT_RUN", "Create step not run because the update step failed.", plan)
            for row in plan.create_rows
        ]
    try:
        create_variants(shopify, plan)
    except RuntimeError as exc:
        return [report_row(row, "UPDATED", "Existing variant updated.", plan) for row in plan.update_rows] + [
            report_row(row, "ERROR", f"Creating variants failed: {exc}", plan) for row in plan.create_rows
        ]
    return [
        report_row(row, "UPDATED" if row.variant_id else "CREATED", "Linked variant operation completed.", plan)
        for row in plan.source_rows
    ]


def write_report(rows: list[dict[str, str]]) -> None:
    REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "excel_row", "product_reference_type", "product_reference", "variant_id", "operation",
        "option1_value", "metaobject_handle", "status", "detail", "sku", "barcode", "price",
        "ignored_populated_columns", "shopify_product_id", "shopify_handle", "shopify_option_id",
        "added_linked_value_ids",
    ]
    with REPORT_PATH.open("w", newline="", encoding="utf-8") as destination:
        writer = csv.DictWriter(destination, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--store", required=True)
    parser.add_argument(
        "--workbook-url",
        required=True,
        help="Direct HTTPS download URL for the unchanged prepared source workbook.",
    )
    parser.add_argument("--mode", choices=("dry-run", "execute"), default="dry-run")
    args = parser.parse_args()

    try:
        download_workbook(args.workbook_url, INPUT_PATH)
        groups = source_groups(INPUT_PATH)
        shopify = Shopify(configured_store(args.store))
        rows: list[dict[str, str]] = []
        for group in groups:
            try:
                plan = plan_group(group, shopify)
                rows.extend(execute_plan(plan, shopify, args.mode))
            except RuntimeError as exc:
                rows.extend(report_row(row, "ERROR", str(exc)) for row in group)
        write_report(rows)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 2

    summary = Counter(row["status"] for row in rows)
    print("Store:", args.store)
    print("Mode:", args.mode)
    print("Products:", len(groups))
    print("Results:", dict(sorted(summary.items())))
    print("Report:", REPORT_PATH)
    return 1 if summary.get("ERROR", 0) else 0


if __name__ == "__main__":
    raise SystemExit(main())
