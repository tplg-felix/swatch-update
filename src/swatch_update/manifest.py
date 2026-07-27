"""Build and load the explicit Color & pattern link manifest."""

from __future__ import annotations

import csv
from collections import defaultdict
from collections.abc import Iterable
from pathlib import Path

from openpyxl import load_workbook

from .handles import normalize_metaobject_handle
from .models import LinkRow

MANIFEST_FIELDNAMES = [
    "Source Excel Row",
    "Product Handle",
    "Product Title",
    "Variant SKU",
    "Option Name",
    "Option Value (Display Name)",
    "Color & Pattern Metaobject Handle",
    "Metaobject Type",
    "Linked Metafield Namespace",
    "Linked Metafield Key",
    "Expected Source Color",
    "Expected Source Material",
    "Post-Import Action",
    "Verification Status",
]

PREFLIGHT_FIELDNAMES = [
    "Metaobject Type",
    "Color & Pattern Metaobject Handle",
    "Expected Display Value",
    "Affected Variant Rows",
    "Source Excel Rows",
    "Preflight Status",
]

REQUIRED_MATRIXIFY_HEADERS = {
    "Handle",
    "Title",
    "Variant SKU",
    "Option 1 Name",
    "Option 1 Value",
}
OPTIONAL_MATRIXIFY_HEADERS = {
    "Variant Metafield: custom.color [single_line_text_field]",
    "Variant Metafield: custom.material [single_line_text_field]",
}


class ManifestError(ValueError):
    """Raised when an import file cannot safely produce a link manifest."""


def _as_text(value: object | None) -> str:
    return "" if value is None else str(value).strip()


def _select_worksheet(workbook, requested_sheet: str):
    if requested_sheet in workbook.sheetnames:
        return workbook[requested_sheet]
    alternatives = [name for name in ("Product", "Products") if name in workbook.sheetnames]
    if len(alternatives) == 1:
        return workbook[alternatives[0]]
    available = ", ".join(workbook.sheetnames)
    raise ManifestError(
        f"Worksheet {requested_sheet!r} was not found. Available worksheets: {available}."
    )


def _header_map(worksheet) -> dict[str, int]:
    headers: dict[str, int] = {}
    for column in range(1, worksheet.max_column + 1):
        value = worksheet.cell(1, column).value
        if value is not None:
            headers[str(value).strip()] = column
    missing = sorted(REQUIRED_MATRIXIFY_HEADERS.difference(headers))
    if missing:
        raise ManifestError(
            "The Matrixify worksheet is missing required column(s): " + ", ".join(missing)
        )
    return headers


def build_manifest(
    input_workbook: Path,
    output_manifest: Path,
    output_preflight: Path | None = None,
    worksheet_name: str = "Product",
    option_name: str = "Color",
) -> list[LinkRow]:
    """Create an audited CSV manifest for slash-containing Color option values.

    The source workbook is opened read-only and never modified. Only values in
    Option 1 whose option name equals ``option_name`` and contains a slash are
    included. The generated metaobject handle is verified later against Shopify.
    """
    if input_workbook.suffix.casefold() not in {".xlsx", ".xlsm"}:
        raise ManifestError("Input must be an .xlsx or .xlsm Matrixify workbook.")
    if not input_workbook.exists():
        raise ManifestError(f"Input workbook was not found: {input_workbook}")

    workbook = load_workbook(input_workbook, read_only=True, data_only=False)
    worksheet = _select_worksheet(workbook, worksheet_name)
    headers = _header_map(worksheet)

    color_column = headers.get("Variant Metafield: custom.color [single_line_text_field]")
    material_column = headers.get("Variant Metafield: custom.material [single_line_text_field]")
    rows: list[LinkRow] = []

    for excel_row in range(2, worksheet.max_row + 1):
        current_option_name = _as_text(worksheet.cell(excel_row, headers["Option 1 Name"]).value)
        current_option_value = _as_text(worksheet.cell(excel_row, headers["Option 1 Value"]).value)
        is_target_option = current_option_name.casefold() == option_name.casefold()
        if not is_target_option or "/" not in current_option_value:
            continue

        product_handle = _as_text(worksheet.cell(excel_row, headers["Handle"]).value)
        if not product_handle:
            raise ManifestError(
                f"Row {excel_row} has a slash-containing Color value but no product Handle."
            )

        rows.append(
            LinkRow(
                source_excel_row=excel_row,
                product_handle=product_handle,
                product_title=_as_text(worksheet.cell(excel_row, headers["Title"]).value),
                variant_sku=_as_text(worksheet.cell(excel_row, headers["Variant SKU"]).value),
                option_name=current_option_name,
                option_value=current_option_value,
                metaobject_handle=normalize_metaobject_handle(current_option_value),
                expected_source_color=(
                    _as_text(worksheet.cell(excel_row, color_column).value) if color_column else ""
                ),
                expected_source_material=(
                    _as_text(worksheet.cell(excel_row, material_column).value)
                    if material_column
                    else ""
                ),
            )
        )

    write_manifest(rows, output_manifest)
    if output_preflight is not None:
        write_preflight(rows, output_preflight)
    return rows


def write_manifest(rows: Iterable[LinkRow], output_path: Path) -> None:
    """Write the execution manifest in a readable, spreadsheet-friendly format."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as destination:
        writer = csv.DictWriter(destination, fieldnames=MANIFEST_FIELDNAMES)
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    "Source Excel Row": row.source_excel_row,
                    "Product Handle": row.product_handle,
                    "Product Title": row.product_title,
                    "Variant SKU": row.variant_sku,
                    "Option Name": row.option_name,
                    "Option Value (Display Name)": row.option_value,
                    "Color & Pattern Metaobject Handle": row.metaobject_handle,
                    "Metaobject Type": "shopify--color-pattern",
                    "Linked Metafield Namespace": "shopify",
                    "Linked Metafield Key": "color-pattern",
                    "Expected Source Color": row.expected_source_color,
                    "Expected Source Material": row.expected_source_material,
                    "Post-Import Action": (
                        "Resolve handle to metaobject GID, then explicitly set linkedMetafieldValue"
                    ),
                    "Verification Status": "VERIFY METAOBJECT EXISTS BEFORE EXECUTION",
                }
            )


def write_preflight(rows: Iterable[LinkRow], output_path: Path) -> None:
    """Write one row per distinct required Color & pattern handle."""
    grouped: dict[str, list[LinkRow]] = defaultdict(list)
    for row in rows:
        grouped[row.metaobject_handle].append(row)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as destination:
        writer = csv.DictWriter(destination, fieldnames=PREFLIGHT_FIELDNAMES)
        writer.writeheader()
        for handle in sorted(grouped, key=str.casefold):
            group = grouped[handle]
            display_values = " | ".join(
                sorted({row.option_value for row in group}, key=str.casefold)
            )
            source_rows = ", ".join(str(row.source_excel_row) for row in group)
            writer.writerow(
                {
                    "Metaobject Type": "shopify--color-pattern",
                    "Color & Pattern Metaobject Handle": handle,
                    "Expected Display Value": display_values,
                    "Affected Variant Rows": len(group),
                    "Source Excel Rows": source_rows,
                    "Preflight Status": "CONFIRM EXISTS AND IS ACTIVE",
                }
            )


def load_manifest(path: Path) -> list[LinkRow]:
    """Read a generated or hand-reviewed manifest before an API run."""
    if not path.exists():
        raise ManifestError(f"Manifest was not found: {path}")
    with path.open(newline="", encoding="utf-8-sig") as source:
        reader = csv.DictReader(source)
        headers = set(reader.fieldnames or [])
        needed = {
            "Source Excel Row",
            "Product Handle",
            "Option Name",
            "Option Value (Display Name)",
            "Color & Pattern Metaobject Handle",
        }
        missing = sorted(needed.difference(headers))
        if missing:
            raise ManifestError("Manifest is missing required column(s): " + ", ".join(missing))
        rows = [LinkRow.from_csv_row(row) for row in reader]
    if not rows:
        raise ManifestError("Manifest contains no link rows.")
    return rows
