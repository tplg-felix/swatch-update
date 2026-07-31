#!/usr/bin/env python3
"""Minimal Matrixify-to-Shopify Color & pattern linker."""

from __future__ import annotations

import argparse
import csv
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

import requests
from openpyxl import load_workbook

API_VERSION = "2026-07"
METAOBJECT_TYPE = "shopify--color-pattern"
LINKED_NAMESPACE = "shopify"
LINKED_KEY = "color-pattern"
MAX_DOWNLOAD_BYTES = 500 * 1024 * 1024

# All configured stores use the same Shopify app credentials.
SHARED_APP = {
    "client_id": "b850c1d033f621b91b9cc45dd96dc137",
    "client_secret": "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",
}

# The GitHub Run workflow "store" dropdown uses these keys.
STORES = {
    "intl": {"shop": "intl-topologie.myshopify.com", **SHARED_APP},
    "hk": {"shop": "theunitstoretw.myshopify.com", **SHARED_APP},
    "tw": {"shop": "wholesale-topologie.myshopify.com", **SHARED_APP},
    "jp": {"shop": "topologiejp.myshopify.com", **SHARED_APP},
    "eu": {"shop": "eu-topologie.myshopify.com", **SHARED_APP},
    "us": {"shop": "topologie-us-web.myshopify.com", **SHARED_APP},
    "kr": {"shop": "topologie-korea.myshopify.com", **SHARED_APP},
    "th": {"shop": "thai-topologie.myshopify.com", **SHARED_APP},
}

PRODUCT_BY_HANDLE_QUERY = """
query ProductByHandle($handle: String!) {
  productByIdentifier(identifier: {handle: $handle}) {
    id
    handle
    options {
      id
      name
      linkedMetafield { namespace key }
      optionValues { id name linkedMetafieldValue }
    }
  }
}
"""

PRODUCT_BY_ID_QUERY = """
query ProductById($id: ID!) {
  product(id: $id) {
    id
    handle
    options {
      id
      name
      linkedMetafield { namespace key }
      optionValues { id name linkedMetafieldValue }
    }
  }
}
"""

METAOBJECT_QUERY = """
query MetaobjectByHandle($type: String!, $handle: String!) {
  metaobjectByHandle(handle: {type: $type, handle: $handle}) {
    id
    handle
  }
}
"""

UPDATE_MUTATION = """
mutation LinkColorOptionValues(
  $productId: ID!,
  $option: OptionUpdateInput!,
  $values: [OptionValueUpdateInput!]
) {
  productOptionUpdate(
    productId: $productId,
    option: $option,
    optionValuesToUpdate: $values
  ) {
    userErrors { field message code }
  }
}
"""


def text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def handle_from_label(label: str) -> str:
    normalized = unicodedata.normalize("NFKD", label).encode("ascii", "ignore").decode()
    normalized = normalized.casefold().replace("&", " and ")
    return re.sub(r"-+", "-", re.sub(r"[^a-z0-9]+", "-", normalized)).strip("-")


def configured_store(store_key: str) -> dict[str, str]:
    key = store_key.strip().casefold()
    if key in STORES:
        return STORES[key]
    raise RuntimeError(f"Unknown store {store_key!r}. Available stores: {', '.join(sorted(STORES))}")


def download_workbook(url: str, destination: Path) -> None:
    parsed = urlparse(url.strip())
    if parsed.scheme != "https" or not parsed.netloc:
        raise RuntimeError("Matrixify URL must be a complete HTTPS URL.")

    try:
        response = requests.get(url, stream=True, timeout=(20, 180), allow_redirects=True)
    except requests.RequestException as exc:
        raise RuntimeError("Could not download the Matrixify workbook.") from exc

    if response.status_code != 200:
        raise RuntimeError(f"Matrixify download returned HTTP {response.status_code}.")
    if int(response.headers.get("Content-Length") or 0) > MAX_DOWNLOAD_BYTES:
        raise RuntimeError("Matrixify workbook exceeds the 500 MB test limit.")

    destination.parent.mkdir(parents=True, exist_ok=True)
    written = 0
    with destination.open("wb") as output:
        for chunk in response.iter_content(chunk_size=1024 * 1024):
            if not chunk:
                continue
            written += len(chunk)
            if written > MAX_DOWNLOAD_BYTES:
                raise RuntimeError("Matrixify workbook exceeds the 500 MB test limit.")
            output.write(chunk)

    if destination.read_bytes()[:4] != b"PK\x03\x04":
        destination.unlink(missing_ok=True)
        raise RuntimeError("Downloaded file is not an Excel workbook.")


def header_column(headers: dict[str, int], *aliases: str) -> int | None:
    for alias in aliases:
        if alias in headers:
            return headers[alias]
    return None


def matrixify_rows(workbook_path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    worksheet = workbook["Product"] if "Product" in workbook.sheetnames else workbook["Products"]
    headers = {
        text(worksheet.cell(1, column).value): column
        for column in range(1, worksheet.max_column + 1)
        if text(worksheet.cell(1, column).value)
    }
    option_name_column = header_column(headers, "Option 1 Name", "Option1 Name")
    option_value_column = header_column(headers, "Option 1 Value", "Option1 Value")
    handle_column = header_column(headers, "Handle")
    id_column = header_column(headers, "ID")
    if not option_name_column or not option_value_column or not (handle_column or id_column):
        raise RuntimeError(
            "Need Option1 Name (or Option 1 Name), Option1 Value (or Option 1 Value), "
            "and either Handle or ID."
        )

    rows: list[dict[str, str]] = []
    previous_reference_type = ""
    previous_reference_value = ""
    for source_row in range(2, worksheet.max_row + 1):
        handle = text(worksheet.cell(source_row, handle_column).value) if handle_column else ""
        product_id = text(worksheet.cell(source_row, id_column).value) if id_column else ""
        if handle:
            previous_reference_type, previous_reference_value = "handle", handle
        elif product_id:
            previous_reference_type, previous_reference_value = "id", product_id

        option_name = text(worksheet.cell(source_row, option_name_column).value)
        option_value = text(worksheet.cell(source_row, option_value_column).value)
        if "/" not in option_value:
            continue
        if not previous_reference_value:
            raise RuntimeError(f"Excel row {source_row} needs a product Handle or ID.")

        rows.append(
            {
                "source_excel_row": str(source_row),
                "product_reference_type": previous_reference_type,
                "product_reference": previous_reference_value,
                "option_name": option_name,
                "option_value": option_value,
                "metaobject_handle": handle_from_label(option_value),
            }
        )

    if not rows:
        raise RuntimeError("No slash-containing Option1 Value values were found.")
    return rows


class Shopify:
    def __init__(self, store: dict[str, str]) -> None:
        self.store = store
        self.token: str | None = None
        self.metaobject_cache: dict[str, dict[str, Any] | None] = {}

    @property
    def base_url(self) -> str:
        return f"https://{self.store['shop']}"

    def access_token(self) -> str:
        if self.token:
            return self.token
        response = requests.post(
            f"{self.base_url}/admin/oauth/access_token",
            data={
                "grant_type": "client_credentials",
                "client_id": self.store["client_id"],
                "client_secret": self.store["client_secret"],
            },
            timeout=30,
        )
        payload = response.json()
        if response.status_code >= 400 or not payload.get("access_token"):
            raise RuntimeError("Shopify token request failed.")
        self.token = payload["access_token"]
        return self.token

    def graphql(self, query: str, variables: dict[str, Any]) -> dict[str, Any]:
        response = requests.post(
            f"{self.base_url}/admin/api/{API_VERSION}/graphql.json",
            json={"query": query, "variables": variables},
            headers={
                "Content-Type": "application/json",
                "X-Shopify-Access-Token": self.access_token(),
            },
            timeout=45,
        )
        payload = response.json()
        if response.status_code >= 400 or payload.get("errors"):
            raise RuntimeError("Shopify GraphQL request failed: " + str(payload))
        return payload.get("data") or {}

    def product(self, reference_type: str, reference: str) -> dict[str, Any] | None:
        if reference_type == "handle":
            return self.graphql(PRODUCT_BY_HANDLE_QUERY, {"handle": reference}).get(
                "productByIdentifier"
            )
        product_id = reference if reference.startswith("gid://") else f"gid://shopify/Product/{reference}"
        return self.graphql(PRODUCT_BY_ID_QUERY, {"id": product_id}).get("product")

    def metaobject(self, handle: str) -> dict[str, Any] | None:
        if handle not in self.metaobject_cache:
            self.metaobject_cache[handle] = self.graphql(
                METAOBJECT_QUERY, {"type": METAOBJECT_TYPE, "handle": handle}
            ).get("metaobjectByHandle")
        return self.metaobject_cache[handle]

    def update(self, product_id: str, option: dict[str, Any], values: list[dict[str, str]]) -> None:
        option_input: dict[str, Any] = {"id": option["id"]}
        if not option.get("linkedMetafield"):
            option_input["linkedMetafield"] = {"namespace": LINKED_NAMESPACE, "key": LINKED_KEY}
        payload = self.graphql(
            UPDATE_MUTATION,
            {"productId": product_id, "option": option_input, "values": values},
        ).get("productOptionUpdate") or {}
        if payload.get("userErrors"):
            raise RuntimeError("Shopify update failed: " + str(payload["userErrors"]))


def find_option(product: dict[str, Any], name: str) -> dict[str, Any] | None:
    matches = [
        option
        for option in product.get("options", [])
        if text(option.get("name")).casefold() == name.casefold()
    ]
    return matches[0] if len(matches) == 1 else None


def result(row: dict[str, str], status: str, detail: str, **ids: str) -> dict[str, str]:
    return {
        **row,
        "status": status,
        "detail": detail,
        "product_id": ids.get("product_id", ""),
        "option_id": ids.get("option_id", ""),
        "option_value_id": ids.get("option_value_id", ""),
        "metaobject_id": ids.get("metaobject_id", ""),
    }


def link(rows: list[dict[str, str]], shopify: Shopify, execute: bool) -> list[dict[str, str]]:
    grouped: dict[tuple[str, str, str], list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        grouped[(row["product_reference_type"], row["product_reference"], row["option_name"])].append(row)

    results: list[dict[str, str]] = []
    for (reference_type, reference, option_name), group in grouped.items():
        try:
            product = shopify.product(reference_type, reference)
        except RuntimeError as exc:
            results.extend(result(row, "ERROR", str(exc)) for row in group)
            continue
        if not product:
            results.extend(result(row, "ERROR", "Product not found.") for row in group)
            continue

        option = find_option(product, option_name)
        if not option:
            results.extend(result(row, "ERROR", "Color option not found.", product_id=product["id"]) for row in group)
            continue
        linked = option.get("linkedMetafield")
        if linked and (linked.get("namespace"), linked.get("key")) != (LINKED_NAMESPACE, LINKED_KEY):
            results.extend(
                result(
                    row,
                    "ERROR",
                    "Color option is linked to a different metafield.",
                    product_id=product["id"],
                    option_id=option["id"],
                )
                for row in group
            )
            continue

        group_results: list[dict[str, str]] = []
        updates: list[dict[str, str]] = []
        failed = False
        for row in group:
            option_values = [
                value for value in option.get("optionValues", []) if value.get("name") == row["option_value"]
            ]
            if len(option_values) != 1:
                group_results.append(
                    result(
                        row,
                        "ERROR",
                        "Expected Color option value was not found exactly once.",
                        product_id=product["id"],
                        option_id=option["id"],
                    )
                )
                failed = True
                continue
            option_value = option_values[0]
            try:
                metaobject = shopify.metaobject(row["metaobject_handle"])
            except RuntimeError as exc:
                group_results.append(
                    result(
                        row,
                        "ERROR",
                        str(exc),
                        product_id=product["id"],
                        option_id=option["id"],
                        option_value_id=option_value["id"],
                    )
                )
                failed = True
                continue
            if not metaobject:
                group_results.append(
                    result(
                        row,
                        "ERROR",
                        "Color & pattern metaobject handle was not found.",
                        product_id=product["id"],
                        option_id=option["id"],
                        option_value_id=option_value["id"],
                    )
                )
                failed = True
                continue

            details = {
                "product_id": product["id"],
                "option_id": option["id"],
                "option_value_id": option_value["id"],
                "metaobject_id": metaobject["id"],
            }
            if option_value.get("linkedMetafieldValue") == metaobject["id"]:
                group_results.append(result(row, "ALREADY_LINKED", "Link already matches.", **details))
            else:
                group_results.append(result(row, "PENDING", "Ready to link.", **details))
                updates.append({"id": option_value["id"], "linkedMetafieldValue": metaobject["id"]})

        if failed:
            for item in group_results:
                if item["status"] == "PENDING":
                    item["status"] = "SKIPPED"
                    item["detail"] = "Skipped because another Color value failed validation."
            results.extend(group_results)
            continue
        if not updates:
            results.extend(group_results)
            continue
        if not execute:
            for item in group_results:
                if item["status"] == "PENDING":
                    item["status"] = "WOULD_LINK"
                    item["detail"] = "Dry run only; no Shopify update sent."
            results.extend(group_results)
            continue
        try:
            shopify.update(product["id"], option, updates)
        except RuntimeError as exc:
            for item in group_results:
                if item["status"] == "PENDING":
                    item["status"] = "ERROR"
                    item["detail"] = str(exc)
        else:
            for item in group_results:
                if item["status"] == "PENDING":
                    item["status"] = "LINKED"
                    item["detail"] = "Explicit Color & pattern link applied."
        results.extend(group_results)
    return results


def write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "source_excel_row",
        "product_reference_type",
        "product_reference",
        "option_name",
        "option_value",
        "metaobject_handle",
        "status",
        "detail",
        "product_id",
        "option_id",
        "option_value_id",
        "metaobject_id",
    ]
    with path.open("w", newline="", encoding="utf-8") as output:
        writer = csv.DictWriter(output, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--store", default="intl")
    parser.add_argument("--url", required=True)
    parser.add_argument("--mode", choices=("dry-run", "execute"), default="dry-run")
    args = parser.parse_args()

    try:
        store = configured_store(args.store)
        download_workbook(args.url, Path("input/matrixify.xlsx"))
        rows = matrixify_rows(Path("input/matrixify.xlsx"))
        results = link(rows, Shopify(store), execute=args.mode == "execute")
        write_csv(Path("reports/swatch_linker_results.csv"), results)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 2

    summary = Counter(item["status"] for item in results)
    print("Store:", args.store)
    print("Results:", dict(sorted(summary.items())))
    print("Report: reports/swatch_linker_results.csv")
    return 1 if summary.get("ERROR", 0) else 0


if __name__ == "__main__":
    raise SystemExit(main())
