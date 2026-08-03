#!/usr/bin/env python3
"""Post-Matrixify all-value relinker for Shopify Color & pattern swatches."""

from __future__ import annotations

import argparse
import csv
import sys
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from swatch_linker import (
    LINKED_KEY,
    LINKED_NAMESPACE,
    Shopify,
    configured_store,
    download_workbook,
    handle_from_label,
    text,
)

PRODUCT_OPTIONS_QUERY = """
query ProductOptionsForRelink($id: ID!) {
  product(id: $id) {
    id
    handle
    options {
      id
      name
      position
      linkedMetafield { namespace key }
      optionValues { id name linkedMetafieldValue }
    }
  }
}
"""

UPDATE_MUTATION = """
mutation RelinkAllOptionValues(
  $productId: ID!,
  $option: OptionUpdateInput!,
  $values: [OptionValueUpdateInput!]
) {
  productOptionUpdate(
    productId: $productId,
    option: $option,
    optionValuesToUpdate: $values
  ) {
    userErrors { field message code }
    product {
      id
      handle
      options {
        id
        name
        linkedMetafield { namespace key }
        optionValues { id name linkedMetafieldValue }
      }
    }
  }
}
"""

WORKBOOK_PATH = Path("input/matrixify.xlsx")
REPORT_PATH = Path("reports/relink_all_swatch_values_results.csv")


def header_column(headers: dict[str, int], *aliases: str) -> int | None:
    for alias in aliases:
        if alias in headers:
            return headers[alias]
    return None


def source_products(workbook_path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    sheet_name = "Product" if "Product" in workbook.sheetnames else "Products"
    if sheet_name not in workbook.sheetnames:
        raise RuntimeError("The Matrixify workbook needs a Product or Products sheet.")
    worksheet = workbook[sheet_name]
    headers = {
        text(worksheet.cell(1, column).value): column
        for column in range(1, worksheet.max_column + 1)
        if text(worksheet.cell(1, column).value)
    }
    id_column = header_column(headers, "ID")
    handle_column = header_column(headers, "Handle")
    option_value_column = header_column(headers, "Option1 Value", "Option 1 Value")
    if not option_value_column or not (id_column or handle_column):
        raise RuntimeError("Need Option1 Value (or Option 1 Value) and either ID or Handle.")

    previous_type = ""
    previous_value = ""
    grouped: dict[tuple[str, str], dict[str, Any]] = {}
    for source_row in range(2, worksheet.max_row + 1):
        product_id = text(worksheet.cell(source_row, id_column).value) if id_column else ""
        handle = text(worksheet.cell(source_row, handle_column).value) if handle_column else ""
        if product_id:
            previous_type, previous_value = "id", product_id
        elif handle:
            previous_type, previous_value = "handle", handle
        option_value = text(worksheet.cell(source_row, option_value_column).value)
        if not option_value:
            continue
        if not previous_value:
            raise RuntimeError(f"Excel row {source_row} needs a product ID or Handle.")
        key = (previous_type, previous_value)
        if key not in grouped:
            grouped[key] = {
                "product_reference_type": previous_type,
                "product_reference": previous_value,
                "source_excel_rows": [],
                "source_option1_values": set(),
            }
        grouped[key]["source_excel_rows"].append(str(source_row))
        grouped[key]["source_option1_values"].add(option_value)

    if not grouped:
        raise RuntimeError("No populated Option1 Value cells were found.")
    return [
        {
            "product_reference_type": item["product_reference_type"],
            "product_reference": item["product_reference"],
            "source_excel_rows": ", ".join(item["source_excel_rows"]),
            "source_option1_values": "; ".join(sorted(item["source_option1_values"])),
        }
        for item in grouped.values()
    ]


def product_with_options(shopify: Shopify, reference_type: str, reference: str) -> dict[str, Any] | None:
    if reference_type == "handle":
        product = shopify.product("handle", reference)
        if not product:
            return None
        product_id = product["id"]
    else:
        product_id = reference if reference.startswith("gid://") else f"gid://shopify/Product/{reference}"
    return shopify.graphql(PRODUCT_OPTIONS_QUERY, {"id": product_id}).get("product")


def first_option(product: dict[str, Any]) -> dict[str, Any] | None:
    options = product.get("options", [])
    if not options:
        return None
    return min(options, key=lambda item: item.get("position", 9999))


def outcome(source: dict[str, str], status: str, detail: str, **extra: str) -> dict[str, str]:
    return {
        "product_reference_type": source["product_reference_type"],
        "product_reference": source["product_reference"],
        "source_excel_rows": source["source_excel_rows"],
        "source_option1_values": source["source_option1_values"],
        "status": status,
        "detail": detail,
        "shopify_product_id": extra.get("shopify_product_id", ""),
        "shopify_handle": extra.get("shopify_handle", ""),
        "option_id": extra.get("option_id", ""),
        "option_name": extra.get("option_name", ""),
        "current_option_values": extra.get("current_option_values", ""),
        "derived_handles": extra.get("derived_handles", ""),
        "missing_handles": extra.get("missing_handles", ""),
    }


def relink(source_rows: list[dict[str, str]], shopify: Shopify, execute: bool) -> list[dict[str, str]]:
    results: list[dict[str, str]] = []
    for source in source_rows:
        try:
            product = product_with_options(
                shopify,
                source["product_reference_type"],
                source["product_reference"],
            )
        except RuntimeError as exc:
            results.append(outcome(source, "ERROR", str(exc)))
            continue
        if not product:
            results.append(outcome(source, "ERROR", "Product not found."))
            continue

        option = first_option(product)
        if not option:
            results.append(
                outcome(
                    source,
                    "ERROR",
                    "Product has no option to relink.",
                    shopify_product_id=product["id"],
                    shopify_handle=text(product.get("handle")),
                )
            )
            continue
        linked = option.get("linkedMetafield") or {}
        if linked and (linked.get("namespace"), linked.get("key")) != (LINKED_NAMESPACE, LINKED_KEY):
            results.append(
                outcome(
                    source,
                    "ERROR",
                    "First option is linked to a different metafield; no update sent.",
                    shopify_product_id=product["id"],
                    shopify_handle=text(product.get("handle")),
                    option_id=option["id"],
                    option_name=text(option.get("name")),
                )
            )
            continue

        option_values = option.get("optionValues", [])
        if not option_values:
            results.append(
                outcome(
                    source,
                    "ERROR",
                    "First option has no values to map.",
                    shopify_product_id=product["id"],
                    shopify_handle=text(product.get("handle")),
                    option_id=option["id"],
                    option_name=text(option.get("name")),
                )
            )
            continue

        current_values = [text(value.get("name")) for value in option_values]
        handles = {value["id"]: handle_from_label(text(value.get("name"))) for value in option_values}
        metaobjects: dict[str, dict[str, Any]] = {}
        missing_handles: list[str] = []
        api_error = ""
        for handle in sorted(set(handles.values())):
            try:
                metaobject = shopify.metaobject(handle)
            except RuntimeError as exc:
                api_error = str(exc)
                break
            if not metaobject:
                missing_handles.append(handle)
            else:
                metaobjects[handle] = metaobject

        details = {
            "shopify_product_id": product["id"],
            "shopify_handle": text(product.get("handle")),
            "option_id": option["id"],
            "option_name": text(option.get("name")),
            "current_option_values": "; ".join(current_values),
            "derived_handles": "; ".join(handles[value["id"]] for value in option_values),
            "missing_handles": "; ".join(missing_handles),
        }
        if api_error:
            results.append(outcome(source, "ERROR", api_error, **details))
            continue
        if missing_handles:
            results.append(
                outcome(
                    source,
                    "MISSING_METAOBJECT",
                    "No Shopify update sent because every existing option value must map before relinking.",
                    **details,
                )
            )
            continue

        updates = [
            {"id": value["id"], "linkedMetafieldValue": metaobjects[handles[value["id"]]]["id"]}
            for value in option_values
        ]
        all_linked = linked.get("namespace") == LINKED_NAMESPACE and linked.get("key") == LINKED_KEY and all(
            value.get("linkedMetafieldValue") == metaobjects[handles[value["id"]]]["id"]
            for value in option_values
        )
        if all_linked:
            results.append(outcome(source, "ALREADY_LINKED", "Every existing option value already matches.", **details))
            continue
        if not execute:
            results.append(
                outcome(
                    source,
                    "WOULD_RELINK",
                    "Dry run only; Shopify will receive one atomic option-and-all-values update.",
                    **details,
                )
            )
            continue

        option_input: dict[str, Any] = {"id": option["id"]}
        if not linked:
            option_input["linkedMetafield"] = {"namespace": LINKED_NAMESPACE, "key": LINKED_KEY}
        try:
            payload = shopify.graphql(
                UPDATE_MUTATION,
                {"productId": product["id"], "option": option_input, "values": updates},
            ).get("productOptionUpdate") or {}
        except RuntimeError as exc:
            results.append(outcome(source, "ERROR", str(exc), **details))
            continue
        errors = payload.get("userErrors") or []
        if errors:
            results.append(outcome(source, "ERROR", str(errors), **details))
        else:
            results.append(
                outcome(
                    source,
                    "RELINKED",
                    "First option and every existing value are now linked to Color & pattern metaobjects.",
                    **details,
                )
            )
    return results


def write_report(rows: list[dict[str, str]]) -> None:
    REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "product_reference_type",
        "product_reference",
        "source_excel_rows",
        "source_option1_values",
        "status",
        "detail",
        "shopify_product_id",
        "shopify_handle",
        "option_id",
        "option_name",
        "current_option_values",
        "derived_handles",
        "missing_handles",
    ]
    with REPORT_PATH.open("w", newline="", encoding="utf-8") as destination:
        writer = csv.DictWriter(destination, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--store", required=True)
    parser.add_argument("--url", required=True, help="Matrixify direct-download URL for the completed text import.")
    parser.add_argument("--mode", choices=("dry-run", "execute"), default="dry-run")
    args = parser.parse_args()
    try:
        download_workbook(args.url, WORKBOOK_PATH)
        products = source_products(WORKBOOK_PATH)
        results = relink(products, Shopify(configured_store(args.store)), execute=args.mode == "execute")
        write_report(results)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 2

    summary = Counter(row["status"] for row in results)
    print("Store:", args.store)
    print("Mode:", args.mode)
    print("Products:", len(products))
    print("Results:", dict(sorted(summary.items())))
    print("Report:", REPORT_PATH)
    return 1 if summary.get("ERROR", 0) or summary.get("MISSING_METAOBJECT", 0) else 0


if __name__ == "__main__":
    raise SystemExit(main())
